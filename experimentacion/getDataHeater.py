from __future__ import annotations

import re
from pathlib import Path
from datetime import datetime, time
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook


# =========================================================
# Utilidades generales
# =========================================================

def norm_text(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def norm_key(x: Any) -> str:
    s = norm_text(x).lower()
    s = s.replace("º", "")
    s = s.replace("%", "pct")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def make_unique_columns(cols: List[Any]) -> List[str]:
    out = []
    seen = {}
    for c in cols:
        base = norm_text(c)
        if base == "":
            base = "UNNAMED"
        n = seen.get(base, 0) + 1
        seen[base] = n
        out.append(base if n == 1 else f"{base}__{n}")
    return out


def coerce_numeric_series(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s.astype(str).str.replace(",", ".", regex=False), errors="coerce")


def extract_file_date(file_name: str) -> Optional[str]:
    m = re.match(r"(\d{8})", Path(file_name).name)
    return m.group(1) if m else None


def yyyyMMdd_to_date_str(yyyymmdd: str) -> str:
    return datetime.strptime(yyyymmdd, "%Y%m%d").strftime("%d/%m/%Y")


def clock_to_seconds(x: Any) -> Optional[float]:
    if x is None:
        return None

    if isinstance(x, datetime):
        return float(x.hour * 3600 + x.minute * 60 + x.second)

    if isinstance(x, time):
        return float(x.hour * 3600 + x.minute * 60 + x.second)

    s = str(x).strip()
    if s == "":
        return None

    s = s.replace(".", ":")
    s = re.sub(r"\s+", "", s)

    for fmt in ["%H:%M:%S", "%H:%M"]:
        try:
            dt = datetime.strptime(s, fmt)
            return float(dt.hour * 3600 + dt.minute * 60 + dt.second)
        except Exception:
            pass

    return None


def seconds_to_hhmmss(seconds: Optional[float]) -> Optional[str]:
    if seconds is None or pd.isna(seconds):
        return None
    s = int(round(float(seconds)))
    h = s // 3600
    m = (s % 3600) // 60
    sec = s % 60
    return f"{h:02d}:{m:02d}:{sec:02d}"


def build_datetime_from_date_and_clock(date_str: str, clock_s: float) -> pd.Timestamp:
    base = pd.to_datetime(date_str, dayfirst=True)
    return base + pd.to_timedelta(clock_s, unit="s")


# =========================================================
# Lectura genérica de hojas
# =========================================================

def worksheet_to_df(ws, header_row_excel: int = 1) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < header_row_excel:
        return pd.DataFrame()

    header = list(rows[header_row_excel - 1])
    header = make_unique_columns(header)

    data = rows[header_row_excel:]
    df = pd.DataFrame(data, columns=header)
    df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
    return df


def find_sheet_by_exact_or_contains(wb, candidates: List[str]) -> Optional[str]:
    sheet_names = wb.sheetnames
    norm_map = {name: norm_key(name) for name in sheet_names}
    cand_norm = [norm_key(c) for c in candidates]

    for real_name, nk in norm_map.items():
        if nk in cand_norm:
            return real_name

    for real_name, nk in norm_map.items():
        if any(c in nk for c in cand_norm):
            return real_name

    return None


def find_process_sheet_name(wb, file_date: str) -> Optional[str]:
    if file_date is None:
        return None

    for s in wb.sheetnames:
        nk = norm_key(s)
        if file_date in nk and "prueba" in nk:
            return s
    return None


# =========================================================
# Lectura de hojas concretas
# =========================================================

def read_notes_sheet(ws) -> pd.DataFrame:
    return worksheet_to_df(ws, header_row_excel=1)


def read_process_sheet(ws, file_date_str: str) -> pd.DataFrame:
    df = worksheet_to_df(ws, header_row_excel=1)

    if df.empty:
        return df

    hora_col = None
    for c in df.columns:
        if norm_key(c.split("__")[0]) == norm_key("Hora"):
            hora_col = c
            break

    if hora_col is None:
        raise ValueError("No se encontró columna 'Hora' en la hoja de proceso.")

    df = df.rename(columns={hora_col: "Hora"}).copy()
    df["clock_s"] = df["Hora"].apply(clock_to_seconds)
    df["Hora_str"] = df["clock_s"].apply(seconds_to_hhmmss)
    df["datetime"] = [
        build_datetime_from_date_and_clock(file_date_str, t) if t is not None else pd.NaT
        for t in df["clock_s"]
    ]

    # Convertir el resto a numérico si se puede
    for c in df.columns:
        if c in ["Hora", "Hora_str", "clock_s", "datetime"]:
            continue
        df[c] = coerce_numeric_series(df[c])

    df = df.dropna(subset=["clock_s"]).reset_index(drop=True)
    return df


def read_avg_sheet(ws, file_date_str: str) -> pd.DataFrame:
    df = worksheet_to_df(ws, header_row_excel=1)

    if df.empty:
        return df

    hora_col = None
    for c in df.columns:
        if norm_key(c.split("__")[0]) == norm_key("Hora"):
            hora_col = c
            break

    if hora_col is None:
        raise ValueError("No se encontró columna 'Hora' en la hoja Promedio T.")

    df = df.rename(columns={hora_col: "Hora"}).copy()
    df["clock_s"] = df["Hora"].apply(clock_to_seconds)
    df["Hora_str"] = df["clock_s"].apply(seconds_to_hhmmss)
    df["datetime"] = [
        build_datetime_from_date_and_clock(file_date_str, t) if t is not None else pd.NaT
        for t in df["clock_s"]
    ]

    # Convertir el resto a numérico si se puede
    for c in df.columns:
        if c in ["Hora", "Hora_str", "clock_s", "datetime"]:
            continue
        df[c] = coerce_numeric_series(df[c])

    df = df.dropna(subset=["clock_s"]).reset_index(drop=True)
    return df


# =========================================================
# Notas temporales
# =========================================================

TIME_NOTE_RE = re.compile(
    r"^\s*(?P<h>\d{1,2})[:.](?P<m>\d{2})(?:[:.](?P<s>\d{2}))?\s*(?P<txt>.*)$"
)


def extract_timed_notes(notes_raw: pd.DataFrame) -> pd.DataFrame:
    rows = []

    if notes_raw is None or notes_raw.empty:
        return pd.DataFrame(columns=["clock_s", "time_str", "note", "raw"])

    for cell in notes_raw.values.ravel():
        if pd.isna(cell):
            continue

        text = str(cell)
        sublines = re.split(r"[\r\n]+", text)

        for line in sublines:
            raw = str(line).strip()
            if not raw:
                continue

            m = TIME_NOTE_RE.match(raw)
            if not m:
                continue

            h = int(m.group("h"))
            minute = int(m.group("m"))
            sec = int(m.group("s") or 0)
            note_txt = m.group("txt").strip()

            clock_s = h * 3600 + minute * 60 + sec
            time_str = f"{h:02d}:{minute:02d}:{sec:02d}"

            rows.append({
                "clock_s": float(clock_s),
                "time_str": time_str,
                "note": note_txt,
                "raw": raw,
            })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("clock_s").drop_duplicates(subset=["clock_s", "raw"]).reset_index(drop=True)

    return df


def build_notes_time_table(notes_raw: pd.DataFrame) -> pd.DataFrame:
    notes_df = extract_timed_notes(notes_raw)

    if notes_df.empty:
        return pd.DataFrame(columns=["event_id", "Time", "Description", "clock_s"])

    notes_df = notes_df.sort_values("clock_s").reset_index(drop=True)
    notes_df["event_id"] = [f"E{i:02d}" for i in range(1, len(notes_df) + 1)]

    out = notes_df[["event_id", "time_str", "note", "clock_s"]].copy()
    out = out.rename(columns={
        "time_str": "Time",
        "note": "Description",
    })

    return out


# =========================================================
# Evento de alimentación
# =========================================================

def find_biomass_feed_event(notes_time: pd.DataFrame,
                            keywords: Optional[List[str]] = None) -> Optional[Dict[str, Any]]:
    if keywords is None:
        keywords = [
            "empezamos a alimentar",
            "empieza alimentación",
            "inicio alimentación",
            "introducimos biomasa",
            "introducimos bioresiduo",
            "alimentamos",
            "alimentar",
            "carga biomasa",
            "empezamos a cargar",
        ]

    if notes_time is None or notes_time.empty:
        return None

    df = notes_time.copy()
    df["desc_norm"] = df["Description"].astype(str).str.lower().str.strip()

    pattern = "|".join(re.escape(k.lower()) for k in keywords)
    hits = df[df["desc_norm"].str.contains(pattern, regex=True, na=False)].copy()

    if hits.empty:
        return None

    hits = hits.sort_values("clock_s").reset_index(drop=True)
    return hits.iloc[0].to_dict()


# =========================================================
# Lectura mínima del caso
# =========================================================

def process_case_for_empty_validation(file_path: str | Path) -> Dict[str, Any]:
    file_path = Path(file_path)

    case = {
        "file_name": file_path.name,
        "ok": False,
        "errors": [],
        "sheet_names": [],
        "notes_raw": pd.DataFrame(),
        "notes_time": pd.DataFrame(),
        "dfs": {
            "process": pd.DataFrame(),
            "avgT": pd.DataFrame(),
        },
        "empty_validation": {},
        "empty_segment": {},
    }

    if not file_path.exists():
        case["errors"].append(f"No existe el archivo: {file_path}")
        return case

    file_date = extract_file_date(file_path.name)
    if file_date is None:
        case["errors"].append("No se pudo extraer la fecha del nombre del archivo.")
        return case

    file_date_str = yyyyMMdd_to_date_str(file_date)

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        case["errors"].append(f"No se pudo abrir el Excel: {e}")
        return case

    case["sheet_names"] = wb.sheetnames

    notes_sheet_name = find_sheet_by_exact_or_contains(wb, ["Notas"])
    avg_sheet_name = find_sheet_by_exact_or_contains(wb, ["Promedio T"])
    process_sheet_name = find_process_sheet_name(wb, file_date)

    if notes_sheet_name is None:
        case["errors"].append("No se encontró la hoja 'Notas'.")
        return case

    if process_sheet_name is None and avg_sheet_name is None:
        case["errors"].append("No se encontró ni hoja de proceso ni hoja 'Promedio T'.")
        return case

    try:
        case["notes_raw"] = read_notes_sheet(wb[notes_sheet_name])
        case["notes_time"] = build_notes_time_table(case["notes_raw"])

        if process_sheet_name is not None:
            case["dfs"]["process"] = read_process_sheet(wb[process_sheet_name], file_date_str)

        if avg_sheet_name is not None:
            case["dfs"]["avgT"] = read_avg_sheet(wb[avg_sheet_name], file_date_str)

    except Exception as e:
        case["errors"].append(str(e))
        return case

    case["ok"] = True
    return case


# =========================================================
# Validación de tramo inicial sin biomasa
# =========================================================

def validate_empty_initial_window(case: Dict[str, Any],
                                  min_pre_feed_min: float = 5.0,
                                  feed_keywords: Optional[List[str]] = None) -> Dict[str, Any]:
    result = {
        "ok_empty": False,
        "reason": None,
        "t_start_s": None,
        "t_start_h": None,
        "t_feed_s": None,
        "t_feed_h": None,
        "dt_pre_feed_min": None,
        "event": None,
    }

    if not case.get("ok", False):
        result["reason"] = "El archivo no pasó la lectura mínima."
        return result

    event = find_biomass_feed_event(case.get("notes_time", pd.DataFrame()), keywords=feed_keywords)
    if event is None:
        result["reason"] = "No se encontró evento de inicio de alimentación."
        return result

    result["event"] = event
    result["t_feed_s"] = float(event["clock_s"])
    result["t_feed_h"] = event["Time"]

    t_candidates = []

    for key in ["process", "avgT"]:
        df = case["dfs"].get(key, pd.DataFrame())
        if not df.empty and "clock_s" in df.columns and df["clock_s"].notna().any():
            t_candidates.append(float(df["clock_s"].dropna().min()))

    if len(t_candidates) == 0:
        result["reason"] = "No hay datos temporales válidos antes de validar."
        return result

    t_start_s = min(t_candidates)
    result["t_start_s"] = t_start_s
    result["t_start_h"] = seconds_to_hhmmss(t_start_s)

    dt_pre_feed_min = (result["t_feed_s"] - t_start_s) / 60.0
    result["dt_pre_feed_min"] = dt_pre_feed_min

    if dt_pre_feed_min <= 0.0:
        result["reason"] = "La prueba empieza con biomasa ya introducida."
        return result

    if dt_pre_feed_min < min_pre_feed_min:
        result["reason"] = (
            f"El tramo previo sin biomasa es demasiado corto: "
            f"{dt_pre_feed_min:.2f} min < {min_pre_feed_min:.2f} min."
        )
        return result

    result["ok_empty"] = True
    result["reason"] = "Caso válido para postprocesar tramo inicial sin biomasa."
    return result


# =========================================================
# Procesado de carpeta
# =========================================================

def process_folder_for_empty_validation(folder: str | Path,
                                        min_pre_feed_min: float = 5.0) -> List[Dict[str, Any]]:
    folder = Path(folder)
    files = sorted(folder.glob("*.xlsx"))

    results = []

    for fp in files:
        case = process_case_for_empty_validation(fp)
        case["empty_validation"] = validate_empty_initial_window(
            case,
            min_pre_feed_min=min_pre_feed_min,
        )
        results.append(case)

    return results


def build_empty_validation_summary(cases: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for case in cases:
        ev = case.get("empty_validation", {})
        rows.append({
            "file_name": case.get("file_name"),
            "ok_read": case.get("ok", False),
            "ok_empty": ev.get("ok_empty", False),
            "reason": ev.get("reason"),
            "t_start_h": ev.get("t_start_h"),
            "t_feed_h": ev.get("t_feed_h"),
            "dt_pre_feed_min": ev.get("dt_pre_feed_min"),
            "event_desc": None if ev.get("event") is None else ev["event"].get("Description"),
            "errors": " | ".join(case.get("errors", [])),
        })

    return pd.DataFrame(rows)


# =========================================================
# Construcción del tramo vacío
# =========================================================

def trim_df_between_times(df: pd.DataFrame,
                          t_start_s: Optional[float] = None,
                          t_end_s: Optional[float] = None) -> pd.DataFrame:
    if df is None or df.empty or "clock_s" not in df.columns:
        return pd.DataFrame()

    out = df.copy()

    if t_start_s is not None:
        out = out[out["clock_s"] >= t_start_s]

    if t_end_s is not None:
        out = out[out["clock_s"] <= t_end_s]

    return out.reset_index(drop=True)


def add_relative_time_minutes(df: pd.DataFrame,
                              t0_s: Optional[float] = None,
                              new_col: str = "x_min") -> pd.DataFrame:
    out = df.copy()

    if out.empty or "clock_s" not in out.columns:
        out[new_col] = np.nan
        return out

    if t0_s is None:
        if out["clock_s"].dropna().empty:
            out[new_col] = np.nan
            return out
        t0_s = float(out["clock_s"].dropna().min())

    out[new_col] = (out["clock_s"] - t0_s) / 60.0
    return out


def build_empty_segment(case: Dict[str, Any]) -> Dict[str, Any]:
    out_case = case.copy()

    if "dfs" not in out_case:
        out_case["empty_segment"] = {
            "ok": False,
            "reason": "El caso no contiene dfs."
        }
        return out_case

    ev = out_case.get("empty_validation", {})
    if not ev.get("ok_empty", False):
        out_case["empty_segment"] = {
            "ok": False,
            "reason": "El caso no pasó la validación de tramo vacío."
        }
        return out_case

    t_start_s = ev.get("t_start_s")
    t_feed_s = ev.get("t_feed_s")

    df_process = out_case["dfs"].get("process", pd.DataFrame())
    df_avgT = out_case["dfs"].get("avgT", pd.DataFrame())

    df_process_empty = trim_df_between_times(df_process, t_start_s=t_start_s, t_end_s=t_feed_s)
    df_avgT_empty = trim_df_between_times(df_avgT, t_start_s=t_start_s, t_end_s=t_feed_s)

    df_process_empty = add_relative_time_minutes(df_process_empty, t0_s=t_start_s, new_col="x_min")
    df_avgT_empty = add_relative_time_minutes(df_avgT_empty, t0_s=t_start_s, new_col="x_min")

    out_case["dfs"]["process_empty"] = df_process_empty
    out_case["dfs"]["avgT_empty"] = df_avgT_empty

    out_case["empty_segment"] = {
        "ok": True,
        "reason": "Tramo vacío construido correctamente.",
        "t_start_s": t_start_s,
        "t_end_s": t_feed_s,
        "t_start_h": seconds_to_hhmmss(t_start_s),
        "t_end_h": seconds_to_hhmmss(t_feed_s),
        "duration_min": None if t_start_s is None or t_feed_s is None else (t_feed_s - t_start_s) / 60.0,
        "n_process": len(df_process_empty),
        "n_avgT": len(df_avgT_empty),
    }

    return out_case


def build_empty_segments_for_cases(cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out = []
    for case in cases:
        out.append(build_empty_segment(case))
    return out


def build_empty_segments_summary(cases: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for i, case in enumerate(cases):
        es = case.get("empty_segment", {})
        ev = case.get("empty_validation", {})

        rows.append({
            "idx": i,
            "file_name": case.get("file_name"),
            "ok_read": case.get("ok", False),
            "ok_empty_validation": ev.get("ok_empty", False),
            "ok_empty_segment": es.get("ok", False),
            "reason": es.get("reason"),
            "t_start_h": es.get("t_start_h"),
            "t_end_h": es.get("t_end_h"),
            "duration_min": es.get("duration_min"),
            "n_process": es.get("n_process"),
            "n_avgT": es.get("n_avgT"),
        })

    return pd.DataFrame(rows)

# =========================================================
# Construcción del tramo previo al inicio de flujo
# =========================================================

def find_first_flow_start(case: Dict[str, Any],
                          flow_candidates: Optional[List[str]] = None,
                          flow_threshold: float = 0.0,
                          use_empty_segment: bool = True) -> Dict[str, Any]:
    """
    Busca el primer instante en que el caudal supera flow_threshold.

    Si use_empty_segment=True:
    - busca en process_empty / avgT_empty

    Si use_empty_segment=False:
    - busca en process / avgT
    """
    if flow_candidates is None:
        flow_candidates = [
            "Caudal_aire",
            "Temp_aire-Ten_QA_0",
            "Temp_aire-Ten_QS_0",
        ]

    out = {
        "ok": False,
        "flow_col": None,
        "flow_source": None,
        "t_flow_start_s": None,
        "t_flow_start_h": None,
        "row_index": None,
        "reason": None,
    }

    if use_empty_segment:
        dfp = case.get("dfs", {}).get("process_empty", pd.DataFrame()).copy()
        dfa = case.get("dfs", {}).get("avgT_empty", pd.DataFrame()).copy()
    else:
        dfp = case.get("dfs", {}).get("process", pd.DataFrame()).copy()
        dfa = case.get("dfs", {}).get("avgT", pd.DataFrame()).copy()

    for c in flow_candidates:
        if not dfp.empty and c in dfp.columns:
            s = pd.to_numeric(dfp[c], errors="coerce")
            mask = s > flow_threshold
            if mask.any():
                idx = int(mask.idxmax())
                t = float(dfp.loc[idx, "clock_s"])
                out["ok"] = True
                out["flow_col"] = c
                out["flow_source"] = "process_empty" if use_empty_segment else "process"
                out["t_flow_start_s"] = t
                out["t_flow_start_h"] = seconds_to_hhmmss(t)
                out["row_index"] = idx
                out["reason"] = "Inicio de flujo detectado en process."
                return out

    for c in flow_candidates:
        if not dfa.empty and c in dfa.columns:
            s = pd.to_numeric(dfa[c], errors="coerce")
            mask = s > flow_threshold
            if mask.any():
                idx = int(mask.idxmax())
                t = float(dfa.loc[idx, "clock_s"])
                out["ok"] = True
                out["flow_col"] = c
                out["flow_source"] = "avgT_empty" if use_empty_segment else "avgT"
                out["t_flow_start_s"] = t
                out["t_flow_start_h"] = seconds_to_hhmmss(t)
                out["row_index"] = idx
                out["reason"] = "Inicio de flujo detectado en avgT."
                return out

    out["reason"] = f"No se encontró ningún instante con caudal > {flow_threshold}."
    return out


def build_pre_flow_segment(case: Dict[str, Any],
                           flow_candidates: Optional[List[str]] = None,
                           flow_threshold: float = 0.0) -> Dict[str, Any]:
    """
    Construye el tramo previo al inicio del flujo:
    desde el inicio del tramo vacío hasta el primer instante con caudal > flow_threshold.

    Añade:
    - case["dfs"]["process_pre_flow"]
    - case["dfs"]["avgT_pre_flow"]
    - case["pre_flow_segment"]
    """
    out_case = case.copy()

    if not out_case.get("empty_segment", {}).get("ok", False):
        out_case["pre_flow_segment"] = {
            "ok": False,
            "reason": "El caso no tiene empty_segment válido."
        }
        return out_case

    flow_info = find_first_flow_start(
        out_case,
        flow_candidates=flow_candidates,
        flow_threshold=flow_threshold,
        use_empty_segment=True,
    )

    if not flow_info["ok"]:
        out_case["pre_flow_segment"] = {
            "ok": False,
            "reason": flow_info["reason"],
            "flow_info": flow_info,
        }
        return out_case

    t_start_s = out_case["empty_segment"]["t_start_s"]
    t_end_s = flow_info["t_flow_start_s"]

    dfp = out_case.get("dfs", {}).get("process_empty", pd.DataFrame())
    dfa = out_case.get("dfs", {}).get("avgT_empty", pd.DataFrame())

    dfp_cut = trim_df_between_times(dfp, t_start_s=t_start_s, t_end_s=t_end_s)
    dfa_cut = trim_df_between_times(dfa, t_start_s=t_start_s, t_end_s=t_end_s)

    dfp_cut = add_relative_time_minutes(dfp_cut, t0_s=t_start_s, new_col="x_min")
    dfa_cut = add_relative_time_minutes(dfa_cut, t0_s=t_start_s, new_col="x_min")

    out_case["dfs"]["process_pre_flow"] = dfp_cut
    out_case["dfs"]["avgT_pre_flow"] = dfa_cut

    out_case["pre_flow_segment"] = {
        "ok": True,
        "reason": "Tramo previo al flujo construido correctamente.",
        "t_start_s": t_start_s,
        "t_end_s": t_end_s,
        "t_start_h": seconds_to_hhmmss(t_start_s),
        "t_end_h": seconds_to_hhmmss(t_end_s),
        "duration_min": (t_end_s - t_start_s) / 60.0 if t_start_s is not None and t_end_s is not None else None,
        "flow_info": flow_info,
        "n_process": len(dfp_cut),
        "n_avgT": len(dfa_cut),
    }

    return out_case


def build_pre_flow_segments_for_cases(cases: List[Dict[str, Any]],
                                      case_indices: Optional[List[int]] = None,
                                      flow_candidates: Optional[List[str]] = None,
                                      flow_threshold: float = 0.0) -> List[Dict[str, Any]]:
    """
    Construye el tramo pre-flow para todos los casos o para una lista de índices concreta.
    """
    out = cases.copy()

    if case_indices is None:
        case_indices = list(range(len(out)))

    for idx in case_indices:
        out[idx] = build_pre_flow_segment(
            out[idx],
            flow_candidates=flow_candidates,
            flow_threshold=flow_threshold,
        )

    return out


def build_pre_flow_summary(cases: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for i, case in enumerate(cases):
        seg = case.get("pre_flow_segment", {})
        flow_info = seg.get("flow_info", {})

        rows.append({
            "idx": i,
            "file_name": case.get("file_name"),
            "ok_pre_flow": seg.get("ok", False),
            "reason": seg.get("reason"),
            "t_start_h": seg.get("t_start_h"),
            "t_end_h": seg.get("t_end_h"),
            "duration_min": seg.get("duration_min"),
            "flow_col": flow_info.get("flow_col"),
            "flow_source": flow_info.get("flow_source"),
        })

    return pd.DataFrame(rows)


# =========================================================
# Detección automática de columnas para plotting
# =========================================================

def detect_empty_plot_columns(case: Dict[str, Any]) -> Dict[str, List[str]]:
    dfp = case.get("dfs", {}).get("process_empty", pd.DataFrame())
    dfa = case.get("dfs", {}).get("avgT_empty", pd.DataFrame())

    proc_cols = list(dfp.columns) if not dfp.empty else []
    avg_cols = list(dfa.columns) if not dfa.empty else []

    avg_temp_cols = [
        c for c in avg_cols
        if c not in ["Hora", "Hora_str", "clock_s", "datetime", "x_min"]
    ]

    proc_temp_cols = [
        c for c in proc_cols
        if (
            ("tg" in norm_key(c) or "ta" in norm_key(c) or "tw" in norm_key(c))
            and c not in ["Hora", "Hora_str", "clock_s", "datetime", "x_min"]
        )
    ]

    flow_cols = [
        c for c in proc_cols + avg_cols
        if (
            "qa" in norm_key(c)
            or "qs" in norm_key(c)
            or "caudal" in norm_key(c)
            or "flow" in norm_key(c)
        )
        and c not in ["Hora", "Hora_str", "clock_s", "datetime", "x_min"]
    ]

    pressure_cols = [
        c for c in proc_cols
        if (
            "pg" in norm_key(c)
            or "pa" in norm_key(c)
            or "pres" in norm_key(c)
        )
        and c not in ["Hora", "Hora_str", "clock_s", "datetime", "x_min"]
    ]

    def unique_keep_order(seq: List[str]) -> List[str]:
        seen = set()
        out = []
        for x in seq:
            if x not in seen:
                out.append(x)
                seen.add(x)
        return out

    return {
        "avg_temp_cols": unique_keep_order(avg_temp_cols),
        "proc_temp_cols": unique_keep_order(proc_temp_cols),
        "flow_cols": unique_keep_order(flow_cols),
        "pressure_cols": unique_keep_order(pressure_cols),
    }


# =========================================================
# Ejes de tiempo
# =========================================================

def add_top_time_axis(ax, t0_s: float):
    secax = ax.secondary_xaxis(
        "top",
        functions=(
            lambda x: x * 60.0 + t0_s,
            lambda x: (x - t0_s) / 60.0,
        ),
    )

    secax.set_xlabel("Hora")
    secax.xaxis.set_major_locator(mticker.MaxNLocator(8))
    secax.xaxis.set_major_formatter(
        mticker.FuncFormatter(
            lambda x, pos: seconds_to_hhmmss(x)[:5] if x is not None else ""
        )
    )
    return secax


# =========================================================
# Plotting
# =========================================================

def plot_empty_temperatures_avg(case: Dict[str, Any],
                                cols: Optional[List[str]] = None,
                                ax=None):
    df = case.get("dfs", {}).get("avgT_empty", pd.DataFrame()).copy()

    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 4))

    if df.empty:
        ax.text(0.5, 0.5, "avgT_empty vacío", transform=ax.transAxes, ha="center", va="center")
        ax.set_ylabel("T avg")
        ax.grid(True, alpha=0.3)
        return ax

    if cols is None:
        cols = detect_empty_plot_columns(case)["avg_temp_cols"]

    plotted = False
    for c in cols:
        if c in df.columns:
            ax.plot(df["x_min"], pd.to_numeric(df[c], errors="coerce"), label=c)
            plotted = True

    if not plotted:
        ax.text(0.5, 0.5, "No hay columnas de temperatura avg detectadas",
                transform=ax.transAxes, ha="center", va="center")

    ax.set_ylabel("T medias [°C]")
    ax.grid(True, alpha=0.3)
    if plotted:
        ax.legend(fontsize=8, ncol=2)

    return ax


def plot_empty_temperatures_process(case: Dict[str, Any],
                                    cols: Optional[List[str]] = None,
                                    ax=None):
    df = case.get("dfs", {}).get("process_empty", pd.DataFrame()).copy()

    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 4))

    if df.empty:
        ax.text(0.5, 0.5, "process_empty vacío", transform=ax.transAxes, ha="center", va="center")
        ax.set_ylabel("T proceso")
        ax.grid(True, alpha=0.3)
        return ax

    if cols is None:
        cols = detect_empty_plot_columns(case)["proc_temp_cols"]

    plotted = False
    for c in cols:
        if c in df.columns:
            ax.plot(df["x_min"], pd.to_numeric(df[c], errors="coerce"), label=c)
            plotted = True

    if not plotted:
        ax.text(0.5, 0.5, "No hay columnas de temperatura de proceso detectadas",
                transform=ax.transAxes, ha="center", va="center")

    ax.set_ylabel("T proceso [°C]")
    ax.grid(True, alpha=0.3)
    if plotted:
        ax.legend(fontsize=8, ncol=2)

    return ax


def plot_empty_flows(case: Dict[str, Any],
                     cols: Optional[List[str]] = None,
                     ax=None):
    dfp = case.get("dfs", {}).get("process_empty", pd.DataFrame()).copy()
    dfa = case.get("dfs", {}).get("avgT_empty", pd.DataFrame()).copy()

    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 4))

    if cols is None:
        cols = detect_empty_plot_columns(case)["flow_cols"]

    plotted = False

    for c in cols:
        if not dfa.empty and c in dfa.columns:
            ax.plot(dfa["x_min"], pd.to_numeric(dfa[c], errors="coerce"), label=c)
            plotted = True
        elif not dfp.empty and c in dfp.columns:
            ax.plot(dfp["x_min"], pd.to_numeric(dfp[c], errors="coerce"), label=c)
            plotted = True

    if not plotted:
        ax.text(0.5, 0.5, "No hay columnas de caudal detectadas",
                transform=ax.transAxes, ha="center", va="center")

    ax.set_ylabel("Caudal")
    ax.grid(True, alpha=0.3)
    if plotted:
        ax.legend(fontsize=8, ncol=2)

    return ax


def plot_empty_pressures(case: Dict[str, Any],
                         cols: Optional[List[str]] = None,
                         ax=None):
    df = case.get("dfs", {}).get("process_empty", pd.DataFrame()).copy()

    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 4))

    if df.empty:
        ax.text(0.5, 0.5, "process_empty vacío", transform=ax.transAxes, ha="center", va="center")
        ax.set_ylabel("Presión")
        ax.grid(True, alpha=0.3)
        return ax

    if cols is None:
        cols = detect_empty_plot_columns(case)["pressure_cols"]

    plotted = False
    for c in cols:
        if c in df.columns:
            ax.plot(df["x_min"], pd.to_numeric(df[c], errors="coerce"), label=c)
            plotted = True

    if not plotted:
        ax.text(0.5, 0.5, "No hay columnas de presión detectadas",
                transform=ax.transAxes, ha="center", va="center")

    ax.set_ylabel("Presión")
    ax.grid(True, alpha=0.3)
    if plotted:
        ax.legend(fontsize=8, ncol=2)

    return ax


def plot_empty_case_overview(case: Dict[str, Any],
                             avg_temp_cols: Optional[List[str]] = None,
                             proc_temp_cols: Optional[List[str]] = None,
                             flow_cols: Optional[List[str]] = None,
                             pressure_cols: Optional[List[str]] = None,
                             figsize: Tuple[float, float] = (12, 13)):
    es = case.get("empty_segment", {})
    if not es.get("ok", False):
        raise ValueError("El caso no tiene tramo vacío construido. Ejecuta build_empty_segment(case).")

    t0_s = es.get("t_start_s")
    duration_min = es.get("duration_min")

    fig, axes = plt.subplots(4, 1, figsize=figsize, sharex=True)
    fig.subplots_adjust(hspace=0.08)

    plot_empty_temperatures_avg(case, cols=avg_temp_cols, ax=axes[0])
    plot_empty_temperatures_process(case, cols=proc_temp_cols, ax=axes[1])
    plot_empty_flows(case, cols=flow_cols, ax=axes[2])
    plot_empty_pressures(case, cols=pressure_cols, ax=axes[3])

    axes[-1].set_xlabel("Tiempo desde inicio del ensayo [min]")

    if t0_s is not None:
        add_top_time_axis(axes[0], t0_s=t0_s)

    if duration_min is not None:
        for ax in axes:
            ax.axvline(duration_min, linestyle="--", linewidth=1.0, alpha=0.35)

        axes[0].text(
            duration_min,
            axes[0].get_ylim()[1],
            "Inicio alimentación",
            rotation=90,
            va="top",
            ha="right",
            fontsize=8,
        )

    if duration_min is not None:
        title = (
            f"{case.get('file_name', 'case')} | "
            f"vacío: {es.get('t_start_h')} - {es.get('t_end_h')} | "
            f"{duration_min:.2f} min"
        )
    else:
        title = case.get("file_name", "case")

    fig.suptitle(title, y=0.995)

    return fig, axes

def plot_selected_cases(cases: List[Dict[str, Any]],
                        case_indices: List[int],
                        figsize_per_row: Tuple[float, float] = (12, 3.2),
                        sharex: bool = True,
                        flow_col_priority: Optional[List[str]] = None,
                        temp_col_map: Optional[Dict[str, List[str]]] = None,
                        data_key_process: str = "process_empty",
                        data_key_avg: str = "avgT_empty"):
    """
    Dibuja varios casos seleccionados en una figura de 1 columna y n filas.

    En cada fila:
    - eje y izquierdo: TA_1, TG_1, TG_2, TG_3, TG_4, TG_5
    - eje y derecho: caudal

    Parameters
    ----------
    cases : list[dict]
        Lista de casos procesados.
    case_indices : list[int]
        Índices de casos a representar.
    figsize_per_row : tuple(float, float)
        Tamaño base por fila.
    sharex : bool
        Si True, comparte eje X.
    flow_col_priority : list[str], optional
        Lista priorizada de nombres posibles para el caudal.
    temp_col_map : dict[str, list[str]], optional
        Diccionario con nombres alternativos de cada sonda.
    data_key_process : str
        Clave del dataframe de proceso a usar.
        Ejemplos:
        - "process_empty"
        - "process_pre_flow"
    data_key_avg : str
        Clave del dataframe avgT a usar.
        Ejemplos:
        - "avgT_empty"
        - "avgT_pre_flow"

    Returns
    -------
    fig, axes
    """
    if flow_col_priority is None:
        flow_col_priority = [
            "Caudal_aire",
            "Temp_aire-Ten_QA_0",
            "Temp_aire-Ten_QS_0",
        ]

    if temp_col_map is None:
        temp_col_map = {
            "TA_1": ["Temp_aire-Ten_TA_1", "TA_1"],
            "TG_1": ["Temp_aire-Ten_TG_1", "TG_1"],
            "TG_2": ["Temp_aire-Ten_TG_2", "TG_2"],
            "TG_3": ["Temp_aire-Ten_TG_3", "TG_3"],
            "TG_4": ["Temp_aire-Ten_TG_4", "TG_4"],
            "TG_5": ["Temp_aire-Ten_TG_5", "TG_5", "Temp_aire-Ten_TG_9"],
        }

    def find_first_existing_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
        for c in candidates:
            if c in df.columns:
                return c
        return None

    n = len(case_indices)
    if n == 0:
        raise ValueError("case_indices está vacío.")

    fig_w = figsize_per_row[0]
    fig_h = figsize_per_row[1] * n

    fig, axes = plt.subplots(n, 1, figsize=(fig_w, fig_h), sharex=sharex)
    if n == 1:
        axes = [axes]

    for ax, idx in zip(axes, case_indices):
        if idx < 0 or idx >= len(cases):
            ax.text(0.5, 0.5, f"Índice fuera de rango: {idx}",
                    transform=ax.transAxes, ha="center", va="center")
            ax.grid(True, alpha=0.3)
            continue

        case = cases[idx]
        df = case.get("dfs", {}).get(data_key_process, pd.DataFrame()).copy()
        dfa = case.get("dfs", {}).get(data_key_avg, pd.DataFrame()).copy()

        if df.empty:
            ax.text(
                0.5,
                0.5,
                f"{case.get('file_name', f'case {idx}')}:\n{data_key_process} vacío",
                transform=ax.transAxes,
                ha="center",
                va="center"
            )
            ax.grid(True, alpha=0.3)
            continue

        if "x_min" not in df.columns:
            t0 = float(df["clock_s"].dropna().min()) if "clock_s" in df.columns and df["clock_s"].notna().any() else 0.0
            df["x_min"] = (df["clock_s"] - t0) / 60.0

        plotted_temp = False

        for label, candidates in temp_col_map.items():
            col = find_first_existing_column(df, candidates)
            if col is not None:
                ax.plot(
                    df["x_min"],
                    pd.to_numeric(df[col], errors="coerce"),
                    label=label,
                )
                plotted_temp = True

        ax.set_ylabel("T [°C]")
        ax.grid(True, alpha=0.3)

        ax2 = ax.twinx()
        plotted_flow = False

        flow_col = find_first_existing_column(df, flow_col_priority)
        flow_source = "process"

        if flow_col is None and not dfa.empty:
            if "x_min" not in dfa.columns:
                t0a = float(dfa["clock_s"].dropna().min()) if "clock_s" in dfa.columns and dfa["clock_s"].notna().any() else 0.0
                dfa["x_min"] = (dfa["clock_s"] - t0a) / 60.0
            flow_col = find_first_existing_column(dfa, flow_col_priority)
            flow_source = "avgT"

        if flow_col is not None:
            if flow_source == "process":
                ax2.plot(
                    df["x_min"],
                    pd.to_numeric(df[flow_col], errors="coerce"),
                    linestyle="--",
                    label="Caudal",
                )
            else:
                ax2.plot(
                    dfa["x_min"],
                    pd.to_numeric(dfa[flow_col], errors="coerce"),
                    linestyle="--",
                    label="Caudal",
                )
            plotted_flow = True

        ax2.set_ylabel("Caudal")

        file_name = case.get("file_name", f"case_{idx}")
        title = f"[{idx}] {file_name}"

        if data_key_process == "process_pre_flow":
            seg = case.get("pre_flow_segment", {})
            if seg.get("ok", False) and seg.get("duration_min") is not None:
                title += f" | pre-flow = {seg['duration_min']:.2f} min"
        elif data_key_process == "process_empty":
            seg = case.get("empty_segment", {})
            if seg.get("ok", False) and seg.get("duration_min") is not None:
                title += f" | empty = {seg['duration_min']:.2f} min"

        ax.set_title(title)

        lines_left, labels_left = ax.get_legend_handles_labels()
        lines_right, labels_right = ax2.get_legend_handles_labels()

        if plotted_temp or plotted_flow:
            ax.legend(
                lines_left + lines_right,
                labels_left + labels_right,
                loc="upper right",
                fontsize=8,
                ncol=4,
            )

    axes[-1].set_xlabel("Tiempo desde inicio [min]")
    fig.tight_layout()

    return fig, axes


# =========================================================
# Guardado de figuras
# =========================================================

def save_empty_case_overview(case: Dict[str, Any],
                             out_dir: str | Path,
                             dpi: int = 150,
                             close_fig: bool = True):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    fig, axes = plot_empty_case_overview(case)

    fname = Path(case.get("file_name", "case")).stem + "_empty_overview.png"
    out_path = out_dir / fname
    fig.savefig(out_path, dpi=dpi, bbox_inches="tight")

    if close_fig:
        plt.close(fig)

    return out_path


def save_empty_case_overviews(cases: List[Dict[str, Any]],
                              out_dir: str | Path,
                              only_valid: bool = True,
                              dpi: int = 150) -> List[Path]:
    paths = []

    for case in cases:
        if only_valid and not case.get("empty_segment", {}).get("ok", False):
            continue

        try:
            p = save_empty_case_overview(case, out_dir=out_dir, dpi=dpi, close_fig=True)
            paths.append(p)
        except Exception as e:
            print(f"Error guardando {case.get('file_name')}: {e}")

    return paths

def compute_bed_zone_power(case: Dict[str, Any],
                           data_key: str = "process_pre_flow",
                           temp_cols: Optional[Dict[str, str]] = None,
                           L_m: float = 0.75,
                           ri_m: float = 0.04,
                           ro_m: float = 0.045,
                           rho_wall_kg_m3: float = 7900.0,
                           cp_wall_J_kg_K: float = 500.0,
                           zone_length_m: Optional[float] = None,
                           smooth_window: int = 9,
                           polyorder: int = 2) -> pd.DataFrame:
    """
    Calcula la potencia térmica absorbida por 3 zonas del lecho,
    suponiendo que cada sonda representa la temperatura media de una zona completa.

    Hipótesis:
    - no hay flujo de N2
    - no hay biomasa
    - cada T_i representa la temperatura media de una zona
    - inicialmente solo se considera la pared metálica de cada zona

    Returns
    -------
    pd.DataFrame con:
    - clock_s
    - x_min
    - T_z1, T_z2, T_z3
    - dTdt_z1, dTdt_z2, dTdt_z3
    - Qdot_z1_W, Qdot_z2_W, Qdot_z3_W
    - Qdot_sum_W
    """
    from scipy.signal import savgol_filter

    if temp_cols is None:
        temp_cols = {
            "T_z1": "Temp_aire-Ten_TG_1",
            "T_z2": "Temp_aire-Ten_TG_2",
            "T_z3": "Temp_aire-Ten_TG_3",
        }

    if zone_length_m is None:
        zone_length_m = L_m / 5.0

    df = case.get("dfs", {}).get(data_key, pd.DataFrame()).copy()
    if df.empty:
        raise ValueError(f"El DataFrame '{data_key}' está vacío o no existe.")

    required = ["clock_s"] + list(temp_cols.values())
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Faltan columnas en '{data_key}': {missing}")

    out = pd.DataFrame()
    out["clock_s"] = pd.to_numeric(df["clock_s"], errors="coerce")

    if "x_min" in df.columns:
        out["x_min"] = pd.to_numeric(df["x_min"], errors="coerce")
    else:
        t0 = float(out["clock_s"].dropna().min())
        out["x_min"] = (out["clock_s"] - t0) / 60.0

    for new_name, old_name in temp_cols.items():
        out[new_name] = pd.to_numeric(df[old_name], errors="coerce")

    out = out.dropna(subset=["clock_s"]).reset_index(drop=True)

    # Geometría de una zona
    V_wall_zone_m3 = np.pi * (ro_m**2 - ri_m**2) * zone_length_m
    m_wall_zone_kg = rho_wall_kg_m3 * V_wall_zone_m3
    C_wall_zone_J_K = m_wall_zone_kg * cp_wall_J_kg_K

    out.attrs["zone_length_m"] = zone_length_m
    out.attrs["V_wall_zone_m3"] = V_wall_zone_m3
    out.attrs["m_wall_zone_kg"] = m_wall_zone_kg
    out.attrs["C_wall_zone_J_K"] = C_wall_zone_J_K

    t_s = out["clock_s"].values.astype(float)
    n = len(out)

    if n < 5:
        raise ValueError("Muy pocos puntos para calcular derivadas suavizadas.")

    if smooth_window >= n:
        smooth_window = n - 1 if (n - 1) % 2 == 1 else n - 2
    if smooth_window < 5:
        smooth_window = 5
    if smooth_window % 2 == 0:
        smooth_window += 1
    if smooth_window >= n:
        smooth_window = n if n % 2 == 1 else n - 1

    dt_mean = float(np.nanmean(np.diff(t_s)))

    for key in temp_cols.keys():
        y = out[key].values.astype(float)

        y_smooth = savgol_filter(
            y,
            window_length=smooth_window,
            polyorder=polyorder,
            mode="interp"
        )

        dydt = savgol_filter(
            y,
            window_length=smooth_window,
            polyorder=polyorder,
            deriv=1,
            delta=dt_mean,
            mode="interp"
        )

        zone_name = key.replace("T_", "")
        out[f"{key}_smooth"] = y_smooth
        out[f"dTdt_{zone_name}"] = dydt
        out[f"Qdot_{zone_name}_W"] = C_wall_zone_J_K * dydt

    q_cols = [f"Qdot_{k.replace('T_', '')}_W" for k in temp_cols.keys()]
    out["Qdot_sum_W"] = out[q_cols].sum(axis=1)

    return out

def plot_bed_zone_power(zone_df: pd.DataFrame,
                        case_label: str = "",
                        figsize: Tuple[float, float] = (12, 8)):
    """
    Grafica temperaturas y potencias por zonas.
    """
    fig, axes = plt.subplots(2, 1, figsize=figsize, sharex=True)
    fig.subplots_adjust(hspace=0.08)

    temp_cols = [
        ("T_z1_smooth", "Zona 1"),
        ("T_z2_smooth", "Zona 2"),
        ("T_z3_smooth", "Zona 3"),
    ]
    for col, label in temp_cols:
        if col in zone_df.columns:
            axes[0].plot(zone_df["x_min"], zone_df[col], label=label)

    axes[0].set_ylabel("T [°C]")
    axes[0].grid(True, alpha=0.3)
    axes[0].legend()

    q_cols = [
        ("Qdot_z1_W", "Q̇ zona 1"),
        ("Qdot_z2_W", "Q̇ zona 2"),
        ("Qdot_z3_W", "Q̇ zona 3"),
        ("Qdot_sum_W", "Q̇ total"),
    ]
    for col, label in q_cols:
        if col in zone_df.columns:
            axes[1].plot(zone_df["x_min"], zone_df[col], label=label)

    axes[1].set_ylabel("Q̇ [W]")
    axes[1].set_xlabel("Tiempo desde inicio [min]")
    axes[1].grid(True, alpha=0.3)
    axes[1].legend()

    title = "Potencia térmica por zonas del lecho"
    if case_label:
        title += f" | {case_label}"
    fig.suptitle(title, y=0.98)

    return fig, axes

