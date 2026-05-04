"""
Script to generate themoLibs.xlsx for the solid materials database.
Run from the project root:
    C:\ProgramData\anaconda3\python.exe materials/solids/_gen_themoLibs.py
"""

import json
import os
import numpy as np
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colors ────────────────────────────────────────────────────────────────────
HDR_BG   = "2E4057"  # dark blue header
HDR_FG   = "FFFFFF"  # white text
SEC_BG   = "4A7C59"  # green section
INP_BG   = "EAF4FB"  # light blue — user input cells
OUT_BG   = "FFF8E1"  # light amber — computed output
REF_BG   = "F3E5F5"  # light purple — reference/literature values
WARN_BG  = "FFEBEE"  # light red — warning


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)


def _border_thin():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row, col, text, width=None, bg=HDR_BG, fg=HDR_FG, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, color=fg, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border_thin()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _data_cell(ws, row, col, value=None, bg="FFFFFF", number_format=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.font = _font(bold=bold)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _border_thin()
    if number_format:
        c.number_format = number_format
    return c


# ── Horner evaluation (pure Python for script use) ───────────────────────────
def poly_deltaT(a, Tref, Tmax, T):
    T_eff = max(Tref, min(Tmax, T))
    dT = T_eff - Tref
    result = 0.0
    for coeff in reversed(a):
        result = result * dT + coeff
    return result


# ── Load database ─────────────────────────────────────────────────────────────
script_dir = os.path.dirname(os.path.abspath(__file__))
db_path = os.path.join(script_dir, "soliddb.txt")

materials = {}
with open(db_path, "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if line:
            rec = json.loads(line)
            materials[rec["id"]] = rec


# ── Generate temperature grid for each material ───────────────────────────────
def make_T_grid(Tref, Tmax, n=16):
    return np.linspace(Tref, Tmax, n)


# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: README
# ══════════════════════════════════════════════════════════════════════════════
ws_read = wb.create_sheet("README")
ws_read.column_dimensions["A"].width = 12
ws_read.column_dimensions["B"].width = 90

readme_rows = [
    ("SOLIDDB — Thermal Property Library for Solid Materials", "", "title"),
    ("", "", ""),
    ("Purpose", "This workbook documents the polynomial thermal property fits stored in soliddb.txt.", "h2"),
    ("", "Each material has its own sheet showing: (1) polynomial coefficients, (2) a temperature", ""),
    ("", "grid with property values evaluated from the polynomial, (3) reference literature values.", ""),
    ("", "", ""),
    ("How to add a new material", "", "h2"),
    ("Step 1", "Copy the Template sheet and rename it with the new material id.", ""),
    ("Step 2", "Enter measured or literature T-property data in the blue INPUT columns.", ""),
    ("Step 3", "Fit polynomial coefficients externally (numpy.polyfit or Excel LINEST).", ""),
    ("Step 4", "Enter the fitted coefficients a0..a7 in the coefficient table.", ""),
    ("Step 5", "Verify the computed values vs. reference in the OUTPUT columns.", ""),
    ("Step 6", "Add a new NDJSON line in soliddb.txt following the existing format.", ""),
    ("", "", ""),
    ("Polynomial basis", "f(T) = a0 + a1*(T-Tref) + a2*(T-Tref)^2 + ... + a7*(T-Tref)^7", "h2"),
    ("", "T is clipped to [Tref, Tmax] before evaluation (no extrapolation).", ""),
    ("", "a0 = value of the property at T = Tref (exact by construction).", ""),
    ("", "", ""),
    ("Properties", "rho  [kg/m3]  — density", "h2"),
    ("", "cp   [J/kg/K] — specific heat capacity (mass basis)", ""),
    ("", "k    [W/m/K]  — thermal conductivity", ""),
    ("", "", ""),
    ("Cell colors", "", "h2"),
    ("Blue",   "User input: experimental data or coefficients to be entered.", "input"),
    ("Amber",  "Computed output: polynomial evaluated at T grid.", "output"),
    ("Purple", "Reference: literature or datasheet values for validation.", "ref"),
    ("", "", ""),
    ("Accuracy note", "Polynomial fits for ceramics (Al2O3, SiC) are approximate (±5-15%)", "warn"),
    ("", "over the full Tmax range due to Debye saturation of cp. Use themoLibs to", ""),
    ("", "refine with higher-degree polynomials or narrower Tmax.", ""),
]

row = 1
for col_a, col_b, style in readme_rows:
    if style == "title":
        c = ws_read.cell(row=row, column=1, value=col_a)
        ws_read.merge_cells(f"A{row}:B{row}")
        c.fill = _fill(HDR_BG)
        c.font = _font(bold=True, color=HDR_FG, size=13)
        c.alignment = Alignment(horizontal="center", vertical="center")
    elif style == "h2":
        ca = ws_read.cell(row=row, column=1, value=col_a)
        ca.fill = _fill(SEC_BG)
        ca.font = _font(bold=True, color="FFFFFF")
        ca.alignment = Alignment(horizontal="left", vertical="center")
        cb = ws_read.cell(row=row, column=2, value=col_b)
        cb.fill = _fill(SEC_BG)
        cb.font = _font(bold=True, color="FFFFFF")
    elif style == "input":
        ca = ws_read.cell(row=row, column=1, value=col_a)
        ca.fill = _fill(INP_BG)
        cb = ws_read.cell(row=row, column=2, value=col_b)
        cb.fill = _fill(INP_BG)
    elif style == "output":
        ca = ws_read.cell(row=row, column=1, value=col_a)
        ca.fill = _fill(OUT_BG)
        cb = ws_read.cell(row=row, column=2, value=col_b)
        cb.fill = _fill(OUT_BG)
    elif style == "ref":
        ca = ws_read.cell(row=row, column=1, value=col_a)
        ca.fill = _fill(REF_BG)
        cb = ws_read.cell(row=row, column=2, value=col_b)
        cb.fill = _fill(REF_BG)
    elif style == "warn":
        ca = ws_read.cell(row=row, column=1, value=col_a)
        ca.fill = _fill(WARN_BG)
        ca.font = _font(bold=True)
        cb = ws_read.cell(row=row, column=2, value=col_b)
        cb.fill = _fill(WARN_BG)
    else:
        ws_read.cell(row=row, column=1, value=col_a)
        ws_read.cell(row=row, column=2, value=col_b)
    row += 1

ws_read.row_dimensions[1].height = 24


# ══════════════════════════════════════════════════════════════════════════════
# Helper: build a material sheet
# ══════════════════════════════════════════════════════════════════════════════
def build_material_sheet(wb, rec):
    mat_id = rec["id"]
    name_en = rec.get("name_en", mat_id)
    category = rec.get("category", "")
    Tref = float(rec["reference"]["Tref_K"])
    Tmax = float(rec["limits"]["Tmax_K"])
    polys = rec["polynomials"]["properties"]

    a_rho = polys["rho"]["a0_to_a7"]
    a_cp  = polys["cp"]["a0_to_a7"]
    a_k   = polys["k"]["a0_to_a7"]

    ws = wb.create_sheet(mat_id)

    # Column widths
    col_widths = [3, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    ws.merge_cells(f"A{row}:M{row}")
    c = ws.cell(row=row, column=1, value=f"{mat_id} — {name_en}  [{category}]")
    c.fill = _fill(HDR_BG)
    c.font = _font(bold=True, color=HDR_FG, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # Metadata row
    ws.cell(row=row, column=2, value="Tref [K]:").font = _font(bold=True)
    _data_cell(ws, row, 3, Tref, INP_BG, "0.0")
    ws.cell(row=row, column=5, value="Tmax [K]:").font = _font(bold=True)
    _data_cell(ws, row, 6, Tmax, INP_BG, "0.0")
    ws.cell(row=row, column=8, value="Category:").font = _font(bold=True)
    ws.cell(row=row, column=9, value=category)
    row += 1
    row += 1  # blank

    # ── Coefficient table ─────────────────────────────────────────────────────
    _header_cell(ws, row, 2, "Coeff", bg=HDR_BG)
    _header_cell(ws, row, 3, "rho [kg/m³]", bg="2E7D32")
    _header_cell(ws, row, 4, "cp [J/kg/K]", bg="1565C0")
    _header_cell(ws, row, 5, "k [W/m/K]",  bg="6A1B9A")
    row += 1

    coeff_names = [f"a{i}" for i in range(8)]
    for i in range(8):
        _data_cell(ws, row, 2, coeff_names[i], bg="F5F5F5", bold=True)
        _data_cell(ws, row, 3, a_rho[i] if i < len(a_rho) else 0.0, INP_BG, "0.00000E+00")
        _data_cell(ws, row, 4, a_cp[i]  if i < len(a_cp)  else 0.0, INP_BG, "0.00000E+00")
        _data_cell(ws, row, 5, a_k[i]   if i < len(a_k)   else 0.0, INP_BG, "0.00000E+00")
        row += 1

    row += 1  # blank

    # ── Temperature grid table ────────────────────────────────────────────────
    _header_cell(ws, row, 2, "T [K]",       bg="37474F", width=14)
    _header_cell(ws, row, 3, "rho_poly\n[kg/m³]",  bg="2E7D32", width=14)
    _header_cell(ws, row, 4, "cp_poly\n[J/kg/K]",  bg="1565C0", width=14)
    _header_cell(ws, row, 5, "k_poly\n[W/m/K]",    bg="6A1B9A", width=14)
    _header_cell(ws, row, 6, "rho_ref\n[kg/m³]",   bg="6A1B9A", width=14)
    _header_cell(ws, row, 7, "cp_ref\n[J/kg/K]",   bg="6A1B9A", width=14)
    _header_cell(ws, row, 8, "k_ref\n[W/m/K]",     bg="6A1B9A", width=14)
    _header_cell(ws, row, 9, "err_rho %", bg="BF360C", width=10)
    _header_cell(ws, row, 10,"err_cp %",  bg="BF360C", width=10)
    _header_cell(ws, row, 11,"err_k %",   bg="BF360C", width=10)
    ws.row_dimensions[row].height = 30
    row += 1

    T_grid = make_T_grid(Tref, Tmax, n=16)
    for T_val in T_grid:
        rho_p = poly_deltaT(a_rho, Tref, Tmax, T_val)
        cp_p  = poly_deltaT(a_cp,  Tref, Tmax, T_val)
        k_p   = poly_deltaT(a_k,   Tref, Tmax, T_val)

        _data_cell(ws, row, 2,  round(T_val, 1),  "F5F5F5", "0.0")
        _data_cell(ws, row, 3,  round(rho_p, 2),  OUT_BG,   "0.00")
        _data_cell(ws, row, 4,  round(cp_p,  2),  OUT_BG,   "0.00")
        _data_cell(ws, row, 5,  round(k_p,   4),  OUT_BG,   "0.0000")
        # Reference cells: blank (user fills in from datasheet)
        _data_cell(ws, row, 6,  None, REF_BG, "0.00")
        _data_cell(ws, row, 7,  None, REF_BG, "0.00")
        _data_cell(ws, row, 8,  None, REF_BG, "0.0000")
        # Error cells: blank until reference is filled
        _data_cell(ws, row, 9,  None, WARN_BG, "0.00")
        _data_cell(ws, row, 10, None, WARN_BG, "0.00")
        _data_cell(ws, row, 11, None, WARN_BG, "0.00")
        row += 1

    # Cap at Tmax row
    cap = rec.get("cap_at_tmax", {})
    row += 1
    _data_cell(ws, row, 2, "Cap at Tmax:", bg="37474F", bold=True)
    ws.cell(row=row, column=2).font = _font(bold=True, color="FFFFFF")
    _data_cell(ws, row, 3, cap.get("rho_kg_per_m3"), OUT_BG, "0.00")
    _data_cell(ws, row, 4, cap.get("cp_J_per_kgK"),  OUT_BG, "0.00")
    _data_cell(ws, row, 5, cap.get("k_W_per_mK"),    OUT_BG, "0.0000")

    row += 2

    # Notes
    note = ws.cell(row=row, column=2,
        value="Fill in the purple 'ref' columns with datasheet values to compute % errors. "
              "Update a0..a7 in soliddb.txt when satisfied with the fit.")
    note.font = _font(bold=False, color="555555", size=9)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"B{row}:K{row}")
    ws.row_dimensions[row].height = 30


# ── Template sheet ────────────────────────────────────────────────────────────
def build_template_sheet(wb):
    ws = wb.create_sheet("Template")
    col_widths = [3, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.merge_cells(f"A{row}:M{row}")
    c = ws.cell(row=row, column=1, value="NEW MATERIAL TEMPLATE — copy this sheet and rename")
    c.fill = _fill("546E7A")
    c.font = _font(bold=True, color="FFFFFF", size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 2

    ws.cell(row=row, column=2, value="Material id:").font = _font(bold=True)
    _data_cell(ws, row, 3, "NEW_ID", INP_BG)
    ws.cell(row=row, column=5, value="Tref [K]:").font = _font(bold=True)
    _data_cell(ws, row, 6, 298.0, INP_BG, "0.0")
    ws.cell(row=row, column=8, value="Tmax [K]:").font = _font(bold=True)
    _data_cell(ws, row, 9, 1273.0, INP_BG, "0.0")
    row += 2

    _header_cell(ws, row, 2, "Coeff", bg=HDR_BG)
    _header_cell(ws, row, 3, "rho [kg/m³]", bg="2E7D32")
    _header_cell(ws, row, 4, "cp [J/kg/K]", bg="1565C0")
    _header_cell(ws, row, 5, "k [W/m/K]",  bg="6A1B9A")
    row += 1

    for i in range(8):
        _data_cell(ws, row, 2, f"a{i}", bg="F5F5F5", bold=True)
        _data_cell(ws, row, 3, 0.0, INP_BG, "0.00000E+00")
        _data_cell(ws, row, 4, 0.0, INP_BG, "0.00000E+00")
        _data_cell(ws, row, 5, 0.0, INP_BG, "0.00000E+00")
        row += 1

    row += 1
    _header_cell(ws, row, 2, "T [K]",      bg="37474F")
    _header_cell(ws, row, 3, "rho_meas\n[kg/m³]",  bg=INP_BG[:-2]+"BB", fg="000000")
    _header_cell(ws, row, 4, "cp_meas\n[J/kg/K]",  bg=INP_BG[:-2]+"BB", fg="000000")
    _header_cell(ws, row, 5, "k_meas\n[W/m/K]",    bg=INP_BG[:-2]+"BB", fg="000000")
    _header_cell(ws, row, 6, "rho_poly\n[kg/m³]",  bg="2E7D32")
    _header_cell(ws, row, 7, "cp_poly\n[J/kg/K]",  bg="1565C0")
    _header_cell(ws, row, 8, "k_poly\n[W/m/K]",    bg="6A1B9A")
    ws.row_dimensions[row].height = 30
    row += 1

    for _ in range(16):
        _data_cell(ws, row, 2, None, INP_BG, "0.0")
        _data_cell(ws, row, 3, None, INP_BG, "0.00")
        _data_cell(ws, row, 4, None, INP_BG, "0.00")
        _data_cell(ws, row, 5, None, INP_BG, "0.0000")
        for col in range(6, 9):
            _data_cell(ws, row, col, None, OUT_BG, "0.0000")
        row += 1

    row += 1
    note = ws.cell(row=row, column=2,
        value="1) Enter T and measured properties (blue). "
              "2) Fit polynomial (numpy.polyfit or Excel LINEST). "
              "3) Enter coefficients a0..a7 above. "
              "4) Verify poly vs. measured in amber columns. "
              "5) Add entry to soliddb.txt.")
    note.font = _font(bold=False, color="555555", size=9)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"B{row}:K{row}")
    ws.row_dimensions[row].height = 45


# ── Build all sheets ──────────────────────────────────────────────────────────
build_template_sheet(wb)
for mat_id, rec in materials.items():
    build_material_sheet(wb, rec)

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(script_dir, "themoLibs.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")
