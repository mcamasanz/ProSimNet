"""
Script to generate themoLibs.xlsx for the gas materials database.
Run from the project root:
    C:\ProgramData\anaconda3\python.exe materials/fluids/_gen_themoLibs.py
"""

import json
import os
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colors ─────────────────────────────────────────────────────────────────────
HDR_BG  = "2E4057"   # dark navy — main headers
HDR_FG  = "FFFFFF"   # white text
SEC_BG  = "1565C0"   # blue — section headers
INP_BG  = "EAF4FB"   # light blue — user input
OUT_BG  = "FFF8E1"   # light amber — polynomial output
REF_BG  = "F3E5F5"   # light purple — reference/literature
WARN_BG = "FFEBEE"   # light red — error / warning

# Per-property header colors
C_MU  = "4A235A"   # dark purple
C_CP  = "1A5276"   # dark blue
C_K   = "1B5E20"   # dark green
C_U   = "7E5109"   # dark amber
C_H   = "78281F"   # dark red
C_S   = "17202A"   # almost black


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)


def _border_thin():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def _hcell(ws, row, col, text, bg=HDR_BG, fg=HDR_FG, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, color=fg, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = _border_thin()
    return c


def _dcell(ws, row, col, value=None, bg="FFFFFF", fmt=None, bold=False, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, color=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _border_thin()
    if fmt:
        c.number_format = fmt
    return c


def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── Horner evaluation ─────────────────────────────────────────────────────────
def poly_eval(a, Tref, Tmax, T):
    T_eff = max(float(Tref), min(float(Tmax), float(T)))
    dT = T_eff - float(Tref)
    result = 0.0
    for coeff in reversed(a):
        result = result * dT + coeff
    return result


# ── Load gasdb ────────────────────────────────────────────────────────────────
script_dir = os.path.dirname(os.path.abspath(__file__))
db_path    = os.path.join(script_dir, "gasdb.txt")

gases = {}       # ordered dict: formula -> record
with open(db_path, "r", encoding="utf-8") as fh:
    for line in fh:
        line = line.strip()
        if line:
            rec = json.loads(line)
            gases[rec["formula"]] = rec

# Property metadata: (key_in_db, display_name, units, format_poly, format_ref, header_color)
PROPS = [
    ("mu", "μ viscosidad",  "Pa·s",    "0.000E+00", "0.000E+00", C_MU),
    ("cp", "cp calor esp.", "J/kg/K",  "0.00",      "0.00",      C_CP),
    ("k",  "k conduct.",    "W/m/K",   "0.0000",    "0.0000",    C_K ),
    ("u",  "u en. interna", "J/mol",   "0.0",       "0.0",       C_U ),
    ("h",  "h entalpía",    "J/mol",   "0.0",       "0.0",       C_H ),
    ("s",  "s entropía",    "J/mol/K", "0.00",      "0.00",      C_S ),
]
N_PROPS = len(PROPS)   # 6


# ══════════════════════════════════════════════════════════════════════════════
# Workbook
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — README
# ══════════════════════════════════════════════════════════════════════════════
ws_r = wb.create_sheet("README")
ws_r.column_dimensions["A"].width = 14
ws_r.column_dimensions["B"].width = 95

readme = [
    ("GASDB — Gas Thermal & Transport Property Library", "", "title"),
    ("", "", ""),
    ("Purpose", "Documents the polynomial fits stored in gasdb.txt for 8 pure gas species.", "h2"),
    ("", "Each species has its own sheet: coefficients, evaluated T-grid, and columns", ""),
    ("", "for entering reference/literature values to validate the polynomial fit.", ""),
    ("", "", ""),
    ("Indexed by", "'formula' key (N2, O2, H2, CH4, C2H6, CO2, CO, H2O).", "h2"),
    ("", "", ""),
    ("How to add a new species", "", "h2"),
    ("Step 1", "Copy the Template sheet and rename it with the chemical formula.", ""),
    ("Step 2", "Enter measured / NIST data in the blue INPUT columns.", ""),
    ("Step 3", "Fit polynomial externally (numpy.polyfit, scipy.curve_fit, or Excel LINEST).", ""),
    ("Step 4", "Enter coefficients a0..a7 in the coefficient table.", ""),
    ("Step 5", "Verify polynomial vs. reference values in the output/error columns.", ""),
    ("Step 6", "Add a new NDJSON line to gasdb.txt following the existing format.", ""),
    ("", "", ""),
    ("Polynomial basis", "f(T) = a0 + a1*(T−Tref) + a2*(T−Tref)² + … + a7*(T−Tref)⁷", "h2"),
    ("", "T is clipped to [Tref, Tmax] before evaluation — no extrapolation.", ""),
    ("", "a0 = value of the property at T = Tref (exact).", ""),
    ("", "", ""),
    ("Properties", "μ  (mu) [Pa·s]     — dynamic viscosity", "h2"),
    ("", "cp         [J/kg·K]  — mass-specific heat capacity", ""),
    ("", "k          [W/m·K]   — thermal conductivity", ""),
    ("", "u          [J/mol]   — molar internal energy  (referenced to Tref)", ""),
    ("", "h          [J/mol]   — molar enthalpy          (referenced to Tref)", ""),
    ("", "s          [J/mol·K] — molar entropy            (referenced to Tref)", ""),
    ("", "", ""),
    ("Transport (LJ)", "sigma_A  [Å]  — Lennard-Jones collision diameter", "h2"),
    ("", "epsilon/k  [K] — Lennard-Jones energy parameter / Boltzmann constant", ""),
    ("", "Used by Chapman-Enskog diffusion coefficient calculation.", ""),
    ("", "", ""),
    ("Cell colors", "", "h2"),
    ("Blue",   "User input: experimental data or coefficients.", "inp"),
    ("Amber",  "Computed: polynomial evaluated at T grid.", "out"),
    ("Purple", "Reference: NIST or datasheet values for validation.", "ref"),
    ("Red",    "Error: % deviation polynomial vs. reference.", "warn"),
    ("", "", ""),
    ("Sources", "NIST WebBook, DIPPR, Perry's Chemical Engineers' Handbook,", "h2"),
    ("", "JANAF Thermochemical Tables, Chapman-Enskog kinetic theory.", ""),
]

r = 1
for col_a, col_b, sty in readme:
    if sty == "title":
        ws_r.merge_cells(f"A{r}:B{r}")
        c = ws_r.cell(row=r, column=1, value=col_a)
        c.fill = _fill(HDR_BG); c.font = _font(True, HDR_FG, 13)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_r.row_dimensions[r].height = 24
    elif sty == "h2":
        ca = ws_r.cell(row=r, column=1, value=col_a)
        ca.fill = _fill(SEC_BG); ca.font = _font(True, "FFFFFF")
        cb = ws_r.cell(row=r, column=2, value=col_b)
        cb.fill = _fill(SEC_BG); cb.font = _font(True, "FFFFFF")
    elif sty == "inp":
        ws_r.cell(row=r, column=1, value=col_a).fill = _fill(INP_BG)
        ws_r.cell(row=r, column=2, value=col_b).fill = _fill(INP_BG)
    elif sty == "out":
        ws_r.cell(row=r, column=1, value=col_a).fill = _fill(OUT_BG)
        ws_r.cell(row=r, column=2, value=col_b).fill = _fill(OUT_BG)
    elif sty == "ref":
        ws_r.cell(row=r, column=1, value=col_a).fill = _fill(REF_BG)
        ws_r.cell(row=r, column=2, value=col_b).fill = _fill(REF_BG)
    elif sty == "warn":
        ws_r.cell(row=r, column=1, value=col_a).fill = _fill(WARN_BG)
        ws_r.cell(row=r, column=2, value=col_b).fill = _fill(WARN_BG)
    else:
        ws_r.cell(row=r, column=1, value=col_a)
        ws_r.cell(row=r, column=2, value=col_b)
    r += 1


# ══════════════════════════════════════════════════════════════════════════════
# Helper: T grid for a species
# ══════════════════════════════════════════════════════════════════════════════
def make_T_grid(Tref, Tmax, n=20):
    return np.linspace(Tref, Tmax, n).tolist()


# ══════════════════════════════════════════════════════════════════════════════
# Build one species sheet
# ══════════════════════════════════════════════════════════════════════════════
def build_gas_sheet(wb, rec):
    formula  = rec["formula"]
    name_es  = rec.get("name_es", formula)
    Tref     = float(rec["reference"]["Tref_K"])
    Tmax     = float(rec["limits"]["Tmax_K"])
    MW       = float(rec["molar_mass"]["value"])
    sigma    = float(rec["transport"]["lj"]["sigma_A"])
    epskB    = float(rec["transport"]["lj"]["epsilon_over_k_K"])
    inchikey = rec.get("inchikey", "")
    cas      = rec.get("cas", "")
    polys    = rec["polynomials"]["properties"]
    cap      = rec.get("cap_at_tmax", {})

    # Column layout:
    # A(1)=label  B(2)=T  C-H(3-8)=poly  I-N(9-14)=ref  O-Q(15-17)=err_mu/cp/k
    # For coeff table: A=coeff  B-G=properties (6 cols)

    ws = wb.create_sheet(formula)

    # Column widths
    col_w = {1:5, 2:11, 3:13, 4:11, 5:10, 6:12, 7:12, 8:10,
              9:13, 10:11, 11:10, 12:12, 13:12, 14:10,
              15:9, 16:9, 17:9}
    for col, w in col_w.items():
        _set_col_width(ws, col, w)

    r = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:Q{r}")
    c = ws.cell(row=r, column=1, value=f"{formula} — {name_es}")
    c.fill = _fill(HDR_BG); c.font = _font(True, HDR_FG, 13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    # ── Metadata row ──────────────────────────────────────────────────────────
    meta = [
        ("Tref [K]:", Tref, "0.0"),
        ("Tmax [K]:", Tmax, "0.0"),
        ("MW [kg/mol]:", MW,  "0.00000E+00"),
        ("σ_LJ [Å]:", sigma, "0.000"),
        ("ε/k_B [K]:", epskB, "0.0"),
    ]
    col = 1
    for label, val, fmt in meta:
        c = ws.cell(row=r, column=col, value=label)
        c.font = _font(bold=True, size=9)
        c.alignment = Alignment(horizontal="right")
        col += 1
        _dcell(ws, r, col, val, INP_BG, fmt)
        col += 1
    r += 1

    if inchikey or cas:
        ws.cell(row=r, column=1, value="InChIKey:").font = _font(bold=True, size=9)
        ws.cell(row=r, column=2, value=inchikey).font = _font(size=9)
        ws.cell(row=r, column=6, value="CAS:").font = _font(bold=True, size=9)
        ws.cell(row=r, column=7, value=cas).font = _font(size=9)
    r += 1
    r += 1  # blank

    # ── Coefficient table ─────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value="COEFICIENTES POLINÓMICOS  f(T) = Σ aᵢ·(T−Tref)ⁱ")
    c.fill = _fill(HDR_BG); c.font = _font(True, HDR_FG, 10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    # Headers for coeff table
    _dcell(ws, r, 1, "Coeff", bg="E8EAF6", bold=True)
    for j, (key, name, units, *_) in enumerate(PROPS):
        _hcell(ws, r, 2+j, f"{name}\n[{units}]", bg=PROPS[j][5], wrap=True)
    ws.row_dimensions[r].height = 32
    r += 1

    for i in range(8):
        _dcell(ws, r, 1, f"a{i}", bg="F5F5F5", bold=True)
        for j, (key, *_) in enumerate(PROPS):
            a_list = polys[key]["a0_to_a7"]
            val = a_list[i] if i < len(a_list) else 0.0
            _dcell(ws, r, 2+j, val, INP_BG, "0.00000E+00")
        r += 1

    r += 1  # blank

    # ── Temperature grid ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:Q{r}")
    c = ws.cell(row=r, column=1, value="EVALUACIÓN EN REJILLA DE TEMPERATURA")
    c.fill = _fill(HDR_BG); c.font = _font(True, HDR_FG, 10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    # Sub-header rows: POLY | REF | ERROR
    # Row 1 of header: merged blocks
    ws.merge_cells(f"B{r}:B{r+1}")
    _hcell(ws, r, 2, "T\n[K]", bg="37474F", wrap=True)

    ws.merge_cells(f"C{r}:H{r}")
    c = ws.cell(row=r, column=3, value="POLINOMIO (calculado)")
    c.fill = _fill("1A5276"); c.font = _font(True, HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"I{r}:N{r}")
    c = ws.cell(row=r, column=9, value="REFERENCIA (literatura / NIST)")
    c.fill = _fill("6A1B9A"); c.font = _font(True, HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"O{r}:Q{r}")
    c = ws.cell(row=r, column=15, value="ERROR %  (poly vs ref)")
    c.fill = _fill("B71C1C"); c.font = _font(True, HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[r].height = 18
    r += 1

    # Row 2 of header: per-property columns
    for j, (key, name, units, *_) in enumerate(PROPS):
        _hcell(ws, r, 3+j,  f"{key}\n[{units}]", bg=PROPS[j][5], wrap=True)
        _hcell(ws, r, 9+j,  f"{key}_ref\n[{units}]", bg="4A148C", wrap=True)
    _hcell(ws, r, 15, "μ %",  bg="C62828", wrap=True)
    _hcell(ws, r, 16, "cp %", bg="C62828", wrap=True)
    _hcell(ws, r, 17, "k %",  bg="C62828", wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1

    # Data rows
    T_grid = make_T_grid(Tref, Tmax, n=20)
    poly_fmts = [p[3] for p in PROPS]
    ref_fmts  = [p[4] for p in PROPS]

    for T_val in T_grid:
        _dcell(ws, r, 2, round(T_val, 1), "F5F5F5", "0.0")
        for j, (key, *_) in enumerate(PROPS):
            a_list = polys[key]["a0_to_a7"]
            val = poly_eval(a_list, Tref, Tmax, T_val)
            _dcell(ws, r, 3+j,  round(val, 8), OUT_BG, poly_fmts[j])
            _dcell(ws, r, 9+j,  None,           REF_BG, ref_fmts[j])
        # Error cells (blank until reference is filled)
        for col in [15, 16, 17]:
            _dcell(ws, r, col, None, WARN_BG, "0.00")
        r += 1

    r += 1

    # ── Cap at Tmax ───────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(row=r, column=1, value=f"Cap at Tmax={Tmax:.0f} K")
    c.fill = _fill("37474F"); c.font = _font(True, HDR_FG, 9)
    c.alignment = Alignment(horizontal="center")

    cap_keys   = ["mu_Pa_s", "cp_J_per_kgK", "k_W_per_mK", "u_J_per_mol", "h_J_per_mol", "s_J_per_molK"]
    cap_fmts   = ["0.000E+00", "0.00", "0.0000", "0.0", "0.0", "0.00"]
    for j, (ckey, cfmt) in enumerate(zip(cap_keys, cap_fmts)):
        _dcell(ws, r, 3+j, cap.get(ckey), OUT_BG, cfmt)

    r += 2

    # ── Notes ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:Q{r}")
    note = ws.cell(row=r, column=1,
        value="Instrucciones: (1) Rellena las columnas REF (púrpura) con datos de NIST/DIPPR. "
              "(2) Calcula error % en las columnas rojas: =(poly−ref)/ref*100. "
              "(3) Ajusta a0..a7 con numpy.polyfit o Excel LINEST hasta minimizar el error. "
              "(4) Actualiza gasdb.txt con los nuevos coeficientes.")
    note.font = _font(size=9, color="555555")
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 36


# ══════════════════════════════════════════════════════════════════════════════
# Template sheet
# ══════════════════════════════════════════════════════════════════════════════
def build_template_sheet(wb):
    ws = wb.create_sheet("Template")

    col_w = {1:5, 2:11, 3:13, 4:11, 5:10, 6:12, 7:12, 8:10,
              9:13, 10:11, 11:10, 12:12, 13:12, 14:10,
              15:9, 16:9, 17:9}
    for col, w in col_w.items():
        _set_col_width(ws, col, w)

    r = 1
    ws.merge_cells(f"A{r}:Q{r}")
    c = ws.cell(row=r, column=1, value="NUEVA ESPECIE — copia esta hoja y renómbrala con la fórmula química")
    c.fill = _fill("546E7A"); c.font = _font(True, HDR_FG, 12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 2

    # Metadata
    meta_labels = ["Fórmula:", "Tref [K]:", "Tmax [K]:", "MW [kg/mol]:", "σ_LJ [Å]:", "ε/k_B [K]:"]
    meta_vals   = ["FORMULA", 298.0, 1000.0, 0.028, 3.5, 100.0]
    meta_fmts   = [None, "0.0", "0.0", "0.00000E+00", "0.000", "0.0"]
    col = 1
    for lbl, val, fmt in zip(meta_labels, meta_vals, meta_fmts):
        c = ws.cell(row=r, column=col, value=lbl)
        c.font = _font(bold=True, size=9); c.alignment = Alignment(horizontal="right")
        col += 1
        _dcell(ws, r, col, val, INP_BG, fmt)
        col += 1
    r += 2

    # Coeff table
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value="COEFICIENTES POLINÓMICOS  f(T) = Σ aᵢ·(T−Tref)ⁱ")
    c.fill = _fill(HDR_BG); c.font = _font(True, HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    _dcell(ws, r, 1, "Coeff", "E8EAF6", bold=True)
    for j, (key, name, units, *_) in enumerate(PROPS):
        _hcell(ws, r, 2+j, f"{name}\n[{units}]", bg=PROPS[j][5], wrap=True)
    ws.row_dimensions[r].height = 32
    r += 1

    for i in range(8):
        _dcell(ws, r, 1, f"a{i}", "F5F5F5", bold=True)
        for j in range(N_PROPS):
            _dcell(ws, r, 2+j, 0.0, INP_BG, "0.00000E+00")
        r += 1

    r += 1

    # T-grid table (blank)
    ws.merge_cells(f"A{r}:Q{r}")
    c = ws.cell(row=r, column=1, value="DATOS Y EVALUACIÓN EN REJILLA DE TEMPERATURA")
    c.fill = _fill(HDR_BG); c.font = _font(True, HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    # Headers (2-row)
    ws.merge_cells(f"B{r}:B{r+1}")
    _hcell(ws, r, 2, "T\n[K]", bg="37474F", wrap=True)

    ws.merge_cells(f"C{r}:H{r}")
    c = ws.cell(row=r, column=3, value="DATOS MEDIDOS / NIST  (entrada)")
    c.fill = _fill("0D47A1"); c.font = _font(True, HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"I{r}:N{r}")
    c = ws.cell(row=r, column=9, value="POLINOMIO (calculado)")
    c.fill = _fill("1A5276"); c.font = _font(True, HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"O{r}:Q{r}")
    c = ws.cell(row=r, column=15, value="ERROR %")
    c.fill = _fill("B71C1C"); c.font = _font(True, HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    for j, (key, name, units, *_) in enumerate(PROPS):
        _hcell(ws, r, 3+j, f"{key}_meas\n[{units}]", bg=PROPS[j][5], wrap=True)
        _hcell(ws, r, 9+j, f"{key}_poly\n[{units}]", bg=PROPS[j][5], wrap=True)
    _hcell(ws, r, 15, "μ %",  bg="C62828")
    _hcell(ws, r, 16, "cp %", bg="C62828")
    _hcell(ws, r, 17, "k %",  bg="C62828")
    ws.row_dimensions[r].height = 28
    r += 1

    for _ in range(20):
        _dcell(ws, r, 2, None, INP_BG, "0.0")
        for j in range(N_PROPS):
            _dcell(ws, r, 3+j,  None, INP_BG, "0.0000")
            _dcell(ws, r, 9+j,  None, OUT_BG, "0.0000")
        for col in [15, 16, 17]:
            _dcell(ws, r, col, None, WARN_BG, "0.00")
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:Q{r}")
    note = ws.cell(row=r, column=1,
        value="(1) Entra T y propiedades medidas en azul. "
              "(2) Ajusta a0..a7 con numpy.polyfit (recomendado grado 7) o Excel LINEST. "
              "(3) Comprueba error % en rojo. "
              "(4) Añade línea NDJSON en gasdb.txt.")
    note.font = _font(size=9, color="555555")
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 36


# ── Assemble workbook ─────────────────────────────────────────────────────────
build_template_sheet(wb)
for formula, rec in gases.items():
    build_gas_sheet(wb, rec)

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(script_dir, "themoLibs.xlsx")
wb.save(out_path)
print(f"Saved:  {out_path}")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
